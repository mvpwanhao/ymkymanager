# Copyright (c) 2026-2027 宛皓 (Wan Hao). All rights reserved.
# 本文件为「云煤矿业产销量管理系统」的组成部分。
# 仅授予云南云煤矿业开发有限公司及其关联方在内部业务系统中使用；
# 未经著作权人书面同意，禁止复制、反编译、转售或二次发行。详见根目录 LICENSE。
"""数据可视化统计引擎：完整复用 report_engine 中的统计口径与计算规则。

统计逻辑一览（与报表/简报一致）：
- 产量 C/D/E：actual_production 台账按生产日期汇总
  C=周产量, D=月累计产量, E=年累计产量
- 销量 F/G/H/I/J/K：actual_sales 台账按周读取
  F=周销量, G=月累计自产煤销量(混合), H=年累计自产煤销量(混合),
  I=掺配煤年累计(合计记录回退), J=外购煤年累计(合计记录回退), K=H+I

G/H 混合逻辑：以补录存储值为基数 + 后续新增 F 值；无存储值时按时间累计
I/J 回退逻辑：从"合计"记录取值，无匹配时回退到最近一期(<=目标周末)
"""

from __future__ import annotations

import io
from datetime import date, timedelta
from typing import Any
from urllib.parse import quote

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.config import get_settings
from app.constants import MINE_LIST, REMOVED_MINE_KEYWORDS
from app.storage import read_records
from app.timeutil import (
    enumerate_weekly_ranges,
    get_26day_month_range,
    get_26day_statistical_month_label,
    get_26day_year_range,
    get_weekly_range,
    today_beijing,
)

# 可视化中使用的煤矿列表（排除双河煤矿，与报表模板一致）
VIZ_MINE_ORDER: tuple[str, ...] = (
    "姚家村",
    "金所",
    "芒东二矿",
    "郭家山",
    "竜浪",
    "胜利",
)

VIZ_MINE_FULL_NAMES: dict[str, str] = {
    "姚家村": "姚家村煤矿",
    "金所": "金所煤矿",
    "芒东二矿": "芒东二矿",
    "郭家山": "郭家山煤矿",
    "竜浪": "竜浪煤矿",
    "胜利": "胜利煤矿",
}


def _exclude_mines(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or "所属煤矿" not in df.columns:
        return df
    mask = ~df["所属煤矿"].astype(str).str.contains("|".join(REMOVED_MINE_KEYWORDS), na=False)
    return df.loc[mask].copy()


def _load_production_df() -> pd.DataFrame:
    """读取产量台账并做类型规范化。"""
    s = get_settings()
    df = _exclude_mines(read_records(s.actual_production_path))
    if df.empty:
        return pd.DataFrame()
    df = df.copy()
    df["生产日期"] = pd.to_datetime(df["生产日期"], errors="coerce").dt.date
    df["所属煤矿"] = df["所属煤矿"].astype(str)
    df["产量(吨)"] = pd.to_numeric(df.get("产量(吨)"), errors="coerce").fillna(0)
    return df


def _load_energy_df() -> pd.DataFrame:
    """读取能源局产销量台账并做类型规范化。"""
    s = get_settings()
    df = _exclude_mines(read_records(s.energy_reporting_path))
    if df.empty or "生产日期" not in df.columns:
        return pd.DataFrame()
    df = df.copy()
    df["生产日期"] = pd.to_datetime(df["生产日期"], errors="coerce").dt.date
    df["所属煤矿"] = df["所属煤矿"].astype(str)
    if "产量(吨)" in df.columns:
        df["产量(吨)"] = pd.to_numeric(df["产量(吨)"], errors="coerce").fillna(0)
    if "销量(吨)" in df.columns:
        df["销量(吨)"] = pd.to_numeric(df["销量(吨)"], errors="coerce").fillna(0)
    return df


def _load_sales_df() -> tuple[pd.DataFrame, pd.DataFrame]:
    """读取销量台账，返回 (矿记录, 合计记录)。"""
    s = get_settings()
    df = read_records(s.actual_sales_path)
    if df.empty:
        return pd.DataFrame(), pd.DataFrame()
    df = df.copy()
    df["周起始日期"] = pd.to_datetime(df["周起始日期"], errors="coerce").dt.date
    df["周结束日期"] = pd.to_datetime(df["周结束日期"], errors="coerce").dt.date
    df["所属煤矿"] = df["所属煤矿"].astype(str)
    df["销量(吨)"] = pd.to_numeric(df["销量(吨)"], errors="coerce").fillna(0)
    for _col in [
        "月累计自产煤销量(吨)",
        "年累计自产煤销量(吨)",
        "年累计掺配煤销量(吨)",
        "年累计外购煤量(吨)",
    ]:
        if _col not in df.columns:
            df[_col] = 0.0
        df[_col] = pd.to_numeric(df[_col], errors="coerce").fillna(0)
    total_mask = df["所属煤矿"] == "合计"
    sales_totals_df = df[total_mask].copy()
    sales_df = _exclude_mines(df[~total_mask])
    return sales_df, sales_totals_df


def _calc_cumul_sales(
    sdf: pd.DataFrame,
    sales_totals_df: pd.DataFrame,
    period_start: date,
    period_end: date,
    cumul_col: str,
) -> float:
    """计算累计自产煤销量（G/H 混合逻辑）。

    与 report_engine 中 _calc_cumul_brief / _calc_cumul 完全一致：
    1. 优先从"合计"记录中取存储值作为基数，加上后续新增 F 值
    2. 无合计记录时，从各矿记录中取存储值作为基数 + 后续 F
    3. 无任何存储值时，按时间累计求和 F
    """
    # 先尝试从"合计"记录取
    if not sales_totals_df.empty:
        period_tot = sales_totals_df[
            (sales_totals_df["周结束日期"] >= period_start)
            & (sales_totals_df["周结束日期"] <= period_end)
        ]
        if not period_tot.empty:
            stored = period_tot[period_tot[cumul_col] > 0]
            if not stored.empty:
                latest = stored.sort_values("周结束日期").iloc[-1]
                base = float(latest[cumul_col])
                base_we = latest["周结束日期"]
                add = sdf[
                    (sdf["周结束日期"] > base_we)
                    & (sdf["周结束日期"] <= period_end)
                ]["销量(吨)"].sum()
                return base + float(add)

    # 从各矿记录取
    if sdf.empty:
        return 0.0
    recs = sdf[
        (sdf["周结束日期"] >= period_start) & (sdf["周结束日期"] <= period_end)
    ]
    if recs.empty:
        return 0.0
    stored_recs = recs[recs[cumul_col] > 0]
    if not stored_recs.empty:
        latest = stored_recs.sort_values("周结束日期").iloc[-1]
        base = float(latest[cumul_col])
        base_we = latest["周结束日期"]
        add = recs[recs["周结束日期"] > base_we]["销量(吨)"].sum()
        return base + float(add)
    return float(recs["销量(吨)"].sum())


def _calc_cumul_sales_by_mine(
    sdf: pd.DataFrame,
    period_start: date,
    period_end: date,
    cumul_col: str,
) -> float:
    """按矿计算累计自产煤销量（G/H 混合逻辑，单矿版）。

    与 report_engine 中 _calc_cumul 完全一致。
    """
    if sdf.empty:
        return 0.0
    recs = sdf[
        (sdf["周结束日期"] >= period_start) & (sdf["周结束日期"] <= period_end)
    ]
    if recs.empty:
        return 0.0
    stored = recs[recs[cumul_col] > 0]
    if not stored.empty:
        latest = stored.sort_values("周结束日期").iloc[-1]
        base = float(latest[cumul_col])
        base_we = latest["周结束日期"]
        add = recs[recs["周结束日期"] > base_we]["销量(吨)"].sum()
        return base + float(add)
    return float(recs["销量(吨)"].sum())


def _calc_blended_purchased(
    sales_totals_df: pd.DataFrame,
    sales_df: pd.DataFrame,
    target_week_end: date,
    effective_week_end: date,
) -> tuple[float, float]:
    """计算掺配煤/外购煤年累计（I/J 回退逻辑）。

    与 report_engine 中 generate_brief_report 的 I/J 逻辑完全一致：
    1. 从"合计"记录取，无匹配时回退到最近一期(<=目标周末)
    2. 仍为0时从各矿记录汇总
    """
    # "合计"记录独立回退
    effective_week_end_totals = target_week_end
    if not sales_totals_df.empty:
        exact_tot = sales_totals_df[sales_totals_df["周结束日期"] == target_week_end]
        if exact_tot.empty:
            recent_tot = sales_totals_df[sales_totals_df["周结束日期"] <= target_week_end]
            if not recent_tot.empty:
                effective_week_end_totals = recent_tot["周结束日期"].max()

    blended = 0.0
    purchased = 0.0
    if not sales_totals_df.empty:
        tot_week = sales_totals_df[
            sales_totals_df["周结束日期"] == effective_week_end_totals
        ]
        if not tot_week.empty:
            blended = float(tot_week["年累计掺配煤销量(吨)"].iloc[0])
            purchased = float(tot_week["年累计外购煤量(吨)"].iloc[0])

    if blended == 0.0 and not sales_df.empty:
        week_mines = sales_df[sales_df["周结束日期"] == effective_week_end]
        if not week_mines.empty:
            blended = float(week_mines["年累计掺配煤销量(吨)"].sum())
    if purchased == 0.0 and not sales_df.empty:
        week_mines = sales_df[sales_df["周结束日期"] == effective_week_end]
        if not week_mines.empty:
            purchased = float(week_mines["年累计外购煤量(吨)"].sum())

    return blended, purchased


def build_viz_data(
    *,
    period: str = "year",
    custom_start: date | None = None,
    custom_end: date | None = None,
    stat_year: int | None = None,
    stat_month: str | None = None,
) -> dict[str, Any]:
    """构建可视化数据，统计口径与报表/简报完全一致。

    Args:
        period: "year" | "month" | "custom"
        custom_start/custom_end: 自定义区间（period="custom"时生效）
        stat_year: 指定统计年（如 2026）
        stat_month: 指定统计月（如 "2026-07"）

    Returns:
        dict 包含 KPI、各矿明细、图表数据、时间范围信息
    """
    today = today_beijing()

    # ── 确定时间区间 ──
    if period == "year":
        if stat_year is not None:
            start_date = date(stat_year - 1, 12, 26)
            end_date = date(stat_year, 12, 25)
        else:
            ys, ye = get_26day_year_range(today)
            start_date, end_date = ys, ye
    elif period == "month":
        if stat_month:
            try:
                sm_y, sm_m = (int(x) for x in stat_month.split("-", 1))
                if sm_m > 1:
                    start_date = date(sm_y, sm_m - 1, 26)
                else:
                    start_date = date(sm_y - 1, 12, 26)
                end_date = date(sm_y, sm_m, 25)
            except (ValueError, IndexError):
                ms, me = get_26day_month_range(today)
                start_date, end_date = ms, me
        else:
            ms, me = get_26day_month_range(today)
            start_date, end_date = ms, me
    else:  # custom
        if custom_start and custom_end:
            start_date, end_date = custom_start, custom_end
        else:
            start_date = today - timedelta(days=30)
            end_date = today
        if start_date > end_date:
            start_date, end_date = end_date, start_date
        # 自定义区间不超过今天
        if end_date > today:
            end_date = today

    # ── 加载数据 ──
    prod_df = _load_production_df()
    sales_df, sales_totals_df = _load_sales_df()
    energy_df = _load_energy_df()

    # 累计计算的截止日期：不超过今天（年度/月度的 end_date 可能在未来）
    effective_end = min(end_date, today)

    # ── 能源局昨日数据（今日报能源局，取昨天数据）──
    yesterday = today - timedelta(days=1)
    energy_yest_prod: dict[str, float] = {}
    energy_yest_sales: dict[str, float] = {}
    if not energy_df.empty:
        ey = energy_df[energy_df["生产日期"] == yesterday]
        if not ey.empty:
            if "产量(吨)" in ey.columns:
                gp = ey.groupby("所属煤矿")["产量(吨)"].sum()
                energy_yest_prod = {k: float(v) for k, v in gp.items()}
            if "销量(吨)" in ey.columns:
                gs = ey.groupby("所属煤矿")["销量(吨)"].sum()
                energy_yest_sales = {k: float(v) for k, v in gs.items()}

    # ── 确定目标周的结束日期 ──
    # 用 effective_end 所在周的周末作为"目标周末"
    week_start, week_end = get_weekly_range(effective_end)

    # 确定有效数据周（与简报逻辑一致）
    effective_week_end = week_end
    if not sales_df.empty:
        exact_week = sales_df[sales_df["周结束日期"] == week_end]
        if exact_week.empty:
            recent = sales_df[sales_df["周结束日期"] <= week_end]
            if not recent.empty:
                effective_week_end = recent["周结束日期"].max()

    # ── 统计月/年区间（用于累计计算）──
    month_start, month_end = get_26day_month_range(effective_end)
    year_start, _ = get_26day_year_range(effective_end)

    # ── 产量统计 ──
    def _sum_prod(d_start: date, d_end: date) -> float:
        if prod_df.empty:
            return 0.0
        mask = (prod_df["生产日期"] >= d_start) & (prod_df["生产日期"] <= d_end)
        return float(prod_df.loc[mask, "产量(吨)"].sum())

    period_prod = _sum_prod(start_date, end_date)
    month_prod = _sum_prod(month_start, effective_end)
    year_prod = _sum_prod(year_start, effective_end)
    week_prod = _sum_prod(week_start, week_end)

    # ── 销量统计 ──
    def _sum_sales(d_start: date, d_end: date) -> float:
        if sales_df.empty:
            return 0.0
        mask = (sales_df["周结束日期"] >= d_start) & (sales_df["周结束日期"] <= d_end)
        return float(sales_df.loc[mask, "销量(吨)"].sum())

    period_sales = _sum_sales(start_date, end_date)
    week_sales = _sum_sales(week_start, week_end)
    month_sales = _calc_cumul_sales(sales_df, sales_totals_df, month_start, effective_end, "月累计自产煤销量(吨)")
    year_sales = _calc_cumul_sales(sales_df, sales_totals_df, year_start, effective_end, "年累计自产煤销量(吨)")

    blended, purchased = _calc_blended_purchased(
        sales_totals_df, sales_df, week_end, effective_week_end
    )
    total_sales_k = year_sales + blended  # K = H + I

    # ── 各矿明细 ──
    mine_details: list[dict[str, Any]] = []
    for prefix in VIZ_MINE_ORDER:
        full_name = VIZ_MINE_FULL_NAMES.get(prefix, prefix)

        # 产量
        if prod_df.empty:
            m_week_prod = m_month_prod = m_year_prod = m_period_prod = 0.0
        else:
            mdf = prod_df[prod_df["所属煤矿"].str.startswith(prefix, na=False)]
            m_week_prod = float(mdf[(mdf["生产日期"] >= week_start) & (mdf["生产日期"] <= week_end)]["产量(吨)"].sum())
            m_month_prod = float(mdf[(mdf["生产日期"] >= month_start) & (mdf["生产日期"] <= effective_end)]["产量(吨)"].sum())
            m_year_prod = float(mdf[(mdf["生产日期"] >= year_start) & (mdf["生产日期"] <= effective_end)]["产量(吨)"].sum())
            m_period_prod = float(mdf[(mdf["生产日期"] >= start_date) & (mdf["生产日期"] <= end_date)]["产量(吨)"].sum())

        # 销量
        if sales_df.empty:
            m_week_sales = m_month_sales = m_year_sales = 0.0
        else:
            sdf = sales_df[sales_df["所属煤矿"].str.startswith(prefix, na=False)]
            m_week_sales = float(sdf[(sdf["周起始日期"] == week_start) & (sdf["周结束日期"] == week_end)]["销量(吨)"].sum())
            m_month_sales = _calc_cumul_sales_by_mine(sdf, month_start, effective_end, "月累计自产煤销量(吨)")
            m_year_sales = _calc_cumul_sales_by_mine(sdf, year_start, effective_end, "年累计自产煤销量(吨)")

        # 能源局昨日数据（今日报能源局）
        m_energy_prod = 0.0
        m_energy_sales = 0.0
        for mine_key, val in energy_yest_prod.items():
            if mine_key.startswith(prefix):
                m_energy_prod = val
                break
        for mine_key, val in energy_yest_sales.items():
            if mine_key.startswith(prefix):
                m_energy_sales = val
                break

        mine_details.append({
            "name": full_name,
            "prefix": prefix,
            "period_prod": round(m_period_prod, 2),
            "week_prod": round(m_week_prod, 2),
            "month_prod": round(m_month_prod, 2),
            "year_prod": round(m_year_prod, 2),
            "week_sales": round(m_week_sales, 2),
            "month_sales": round(m_month_sales, 2),
            "year_sales": round(m_year_sales, 2),
            "year_prod_wan": round(m_year_prod / 10000, 1),
            "energy_prod": round(m_energy_prod, 2),
            "energy_sales": round(m_energy_sales, 2),
        })

    # ── 产量趋势（按日）──
    prod_trend: list[dict[str, Any]] = []
    if not prod_df.empty:
        trend_df = prod_df[
            (prod_df["生产日期"] >= start_date) & (prod_df["生产日期"] <= end_date)
        ].copy()
        if "备注" in trend_df.columns:
            trend_df["备注"] = trend_df["备注"].astype(str)
            trend_df = trend_df[~trend_df["备注"].str.contains("补录|年初至", na=False)]
        if not trend_df.empty:
            daily = trend_df.groupby("生产日期")["产量(吨)"].sum().sort_index()
            for d, v in daily.items():
                prod_trend.append({"date": d.isoformat(), "value": round(float(v), 2)})

    # ── 各矿产量趋势（按日）──
    mine_prod_trend: dict[str, list[dict[str, Any]]] = {}
    if not prod_df.empty:
        trend_df = prod_df[
            (prod_df["生产日期"] >= start_date) & (prod_df["生产日期"] <= end_date)
        ].copy()
        if "备注" in trend_df.columns:
            trend_df["备注"] = trend_df["备注"].astype(str)
            trend_df = trend_df[~trend_df["备注"].str.contains("补录|年初至", na=False)]
        if not trend_df.empty:
            for prefix in VIZ_MINE_ORDER:
                full_name = VIZ_MINE_FULL_NAMES.get(prefix, prefix)
                mdf = trend_df[trend_df["所属煤矿"].str.startswith(prefix, na=False)]
                if mdf.empty:
                    mine_prod_trend[full_name] = []
                    continue
                daily = mdf.groupby("生产日期")["产量(吨)"].sum().sort_index()
                mine_prod_trend[full_name] = [
                    {"date": d.isoformat(), "value": round(float(v), 2)} for d, v in daily.items()
                ]

    # ── 销量趋势（按周）──
    sales_trend: list[dict[str, Any]] = []
    mine_sales_trend: dict[str, list[dict[str, Any]]] = {}
    if not sales_df.empty:
        trend_sdf = sales_df[
            (sales_df["周结束日期"] >= start_date) & (sales_df["周结束日期"] <= end_date)
        ].copy()
        if not trend_sdf.empty:
            weekly = trend_sdf.groupby("周结束日期")["销量(吨)"].sum().sort_index()
            for d, v in weekly.items():
                sales_trend.append({"date": d.isoformat(), "value": round(float(v), 2)})
            for prefix in VIZ_MINE_ORDER:
                full_name = VIZ_MINE_FULL_NAMES.get(prefix, prefix)
                mdf = trend_sdf[trend_sdf["所属煤矿"].str.startswith(prefix, na=False)]
                if mdf.empty:
                    mine_sales_trend[full_name] = []
                    continue
                weekly_m = mdf.groupby("周结束日期")["销量(吨)"].sum().sort_index()
                mine_sales_trend[full_name] = [
                    {"date": d.isoformat(), "value": round(float(v), 2)} for d, v in weekly_m.items()
                ]

    # ── 能源局合计 ──
    tot_energy_prod = sum(v for v in energy_yest_prod.values())
    tot_energy_sales = sum(v for v in energy_yest_sales.values())

    # ── KPI ──
    kpis = [
        {"label": "期间产量", "value": round(period_prod, 2), "unit": "吨"},
        {"label": "期间销量", "value": round(period_sales, 2), "unit": "吨"},
        {"label": "月累计自产煤销量", "value": round(month_sales, 2), "unit": "吨"},
        {"label": "年累计自产煤销量", "value": round(year_sales, 2), "unit": "吨"},
        {"label": "掺配煤年累计", "value": round(blended, 2), "unit": "吨"},
        {"label": "外购煤年累计", "value": round(purchased, 2), "unit": "吨"},
        {"label": "合计销售煤量(K=H+I)", "value": round(total_sales_k, 2), "unit": "吨"},
        {"label": "今日报能源局产量", "value": round(tot_energy_prod, 2), "unit": "吨"},
        {"label": "今日报能源局销量", "value": round(tot_energy_sales, 2), "unit": "吨"},
    ]

    # ── 时间范围标签 ──
    period_label = f"{start_date.isoformat()} 至 {end_date.isoformat()}"

    # ── 可选统计年/月列表 ──
    year_options: list[dict[str, Any]] = []
    month_options: list[dict[str, Any]] = []
    if not prod_df.empty:
        min_d = prod_df["生产日期"].min()
        max_d = max(prod_df["生产日期"].max(), today)
        # 统计年
        y_lo = get_26day_year_range(min_d)[1].year
        y_hi = get_26day_year_range(max_d)[1].year
        for y in range(y_hi, y_lo - 1, -1):
            year_options.append({"value": y, "label": f"{y}年（{y-1}.12.26-{y}.12.25）"})
        # 统计月
        e_lo = get_26day_month_range(min_d)[1]
        e_hi = get_26day_month_range(max_d)[1]
        y, m = e_lo.year, e_lo.month
        pairs = []
        while (y, m) <= (e_hi.year, e_hi.month):
            pairs.append((y, m))
            if m == 12:
                y, m = y + 1, 1
            else:
                m += 1
        for y, m in reversed(pairs):
            pm = m - 1 if m > 1 else 12
            month_options.append({"value": f"{y:04d}-{m:02d}", "label": f"{y}年{m}月（{pm}.26-{m}.25）"})

    # 当前选中的统计年/月
    cur_year = get_26day_year_range(end_date)[1].year
    cur_month_end = get_26day_month_range(end_date)[1]
    cur_month_val = f"{cur_month_end.year:04d}-{cur_month_end.month:02d}"

    selected_year = stat_year if (stat_year and year_options) else cur_year
    selected_month = stat_month if (stat_month and month_options) else cur_month_val

    return {
        "period": period,
        "period_label": period_label,
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "week_start": week_start.isoformat(),
        "week_end": week_end.isoformat(),
        "yesterday": yesterday.isoformat(),
        "kpis": kpis,
        "mine_details": mine_details,
        "prod_trend": prod_trend,
        "mine_prod_trend": mine_prod_trend,
        "sales_trend": sales_trend,
        "mine_sales_trend": mine_sales_trend,
        "year_options": year_options,
        "month_options": month_options,
        "selected_year": selected_year,
        "selected_month": selected_month,
        "today": today.isoformat(),
        "stat_month_label": get_26day_statistical_month_label(end_date),
    }


def export_viz_excel(
    *,
    period: str = "year",
    custom_start: date | None = None,
    custom_end: date | None = None,
    stat_year: int | None = None,
    stat_month: str | None = None,
) -> tuple[bytes, str, str]:
    """导出可视化数据为 Excel，返回 (bytes, ascii_name, utf8_name)。"""
    data = build_viz_data(
        period=period,
        custom_start=custom_start,
        custom_end=custom_end,
        stat_year=stat_year,
        stat_month=stat_month,
    )
    md = data["mine_details"]
    s0 = data["start_date"]
    s1 = data["end_date"]

    wb = Workbook()

    # ── Sheet 1: 各矿产销量明细 ──
    ws1 = wb.active
    ws1.title = "各矿产销量明细"
    ws1["A1"] = "云煤矿业产销量统计导出"
    ws1["A1"].font = Font(bold=True, size=14)
    ws1["A2"] = f"统计区间：{s0} 至 {s1}"
    ws1["A2"].font = Font(size=11)
    if data.get("yesterday"):
        ws1["A3"] = f"能源局数据为 {data['yesterday']} 当天数据"
        ws1["A3"].font = Font(size=10, italic=True)

    headers = [
        "煤矿名称", "期间产量(吨)", "周产量(吨)", "月累计产量(吨)", "年累计产量(吨)",
        "周销量(吨)", "月累计自产煤销量(吨)", "年累计自产煤销量(吨)",
        "今日报能源局产量(吨)", "今日报能源局销量(吨)",
    ]
    header_row = 5
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    for c, h in enumerate(headers, start=1):
        cell = ws1.cell(row=header_row, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r, m in enumerate(md, start=header_row + 1):
        vals = [
            m["name"], m["period_prod"], m["week_prod"], m["month_prod"], m["year_prod"],
            m["week_sales"], m["month_sales"], m["year_sales"],
            m["energy_prod"], m["energy_sales"],
        ]
        for c, v in enumerate(vals, start=1):
            ws1.cell(row=r, column=c, value=v)

    # 合计行
    tot_row = header_row + len(md) + 1
    ws1.cell(row=tot_row, column=1, value="合计").font = Font(bold=True)
    for c in range(2, len(headers) + 1):
        col_vals = [m[headers[c - 1].replace("煤矿名称", "name")
                       .replace("期间产量(吨)", "period_prod")
                       .replace("周产量(吨)", "week_prod")
                       .replace("月累计产量(吨)", "month_prod")
                       .replace("年累计产量(吨)", "year_prod")
                       .replace("周销量(吨)", "week_sales")
                       .replace("月累计自产煤销量(吨)", "month_sales")
                       .replace("年累计自产煤销量(吨)", "year_sales")
                       .replace("今日报能源局产量(吨)", "energy_prod")
                       .replace("今日报能源局销量(吨)", "energy_sales")] for m in md]
        ws1.cell(row=tot_row, column=c, value=round(sum(col_vals), 2)).font = Font(bold=True)

    for c in range(1, len(headers) + 1):
        ws1.column_dimensions[get_column_letter(c)].width = 20

    # ── Sheet 2: 日产量明细 ──
    ws2 = wb.create_sheet("日产量明细")
    ws2["A1"] = f"日产量明细（{s0} 至 {s1}）"
    ws2["A1"].font = Font(bold=True, size=13)

    prod_df = _load_production_df()
    if not prod_df.empty:
        pdf = prod_df[
            (prod_df["生产日期"] >= pd.Timestamp(s0).date())
            & (prod_df["生产日期"] <= pd.Timestamp(s1).date())
        ].copy()
        if "备注" in pdf.columns:
            pdf["备注"] = pdf["备注"].astype(str)
            pdf = pdf[~pdf["备注"].str.contains("补录|年初至", na=False)]
        if not pdf.empty:
            g = pdf.groupby(["生产日期", "所属煤矿"], as_index=False)["产量(吨)"].sum()
            wide = g.pivot(index="生产日期", columns="所属煤矿", values="产量(吨)").fillna(0.0)
            for m in MINE_LIST:
                if m not in wide.columns:
                    wide[m] = 0.0
            wide = wide.reindex(columns=[m for m in MINE_LIST], fill_value=0.0).sort_index()
            wide["日合计"] = wide.sum(axis=1)
            for c in wide.columns:
                wide[c] = pd.to_numeric(wide[c], errors="coerce").fillna(0).round(2)
            out = wide.reset_index()
            out["生产日期"] = out["生产日期"].astype(str)
            tot: dict[str, Any] = {"生产日期": "期间合计"}
            for c in wide.columns:
                tot[c] = round(float(wide[c].sum()), 2)
            out = pd.concat([out, pd.DataFrame([tot])], ignore_index=True)

            hdr_row2 = 3
            for c, h in enumerate(list(out.columns), start=1):
                cell = ws2.cell(row=hdr_row2, column=c, value=str(h))
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal="center")
            for r, (_, row) in enumerate(out.iterrows(), start=hdr_row2 + 1):
                for c, v in enumerate(row, start=1):
                    ws2.cell(row=r, column=c, value=v)
            for c in range(1, len(out.columns) + 1):
                ws2.column_dimensions[get_column_letter(c)].width = 16

    # ── Sheet 3: 周销量明细 ──
    ws3 = wb.create_sheet("周销量明细")
    ws3["A1"] = f"周销量明细（{s0} 至 {s1}）"
    ws3["A1"].font = Font(bold=True, size=13)

    sales_df, _ = _load_sales_df()
    if not sales_df.empty:
        sdf = sales_df[
            (sales_df["周结束日期"] >= pd.Timestamp(s0).date())
            & (sales_df["周结束日期"] <= pd.Timestamp(s1).date())
        ].copy()
        if not sdf.empty:
            g2 = sdf.groupby(["周结束日期", "所属煤矿"], as_index=False)["销量(吨)"].sum()
            wide2 = g2.pivot(index="周结束日期", columns="所属煤矿", values="销量(吨)").fillna(0.0)
            for m in MINE_LIST:
                if m not in wide2.columns:
                    wide2[m] = 0.0
            wide2 = wide2.reindex(columns=[m for m in MINE_LIST], fill_value=0.0).sort_index()
            wide2["周合计"] = wide2.sum(axis=1)
            for c in wide2.columns:
                wide2[c] = pd.to_numeric(wide2[c], errors="coerce").fillna(0).round(2)
            out2 = wide2.reset_index()
            out2["周结束日期"] = out2["周结束日期"].astype(str)
            tot2: dict[str, Any] = {"周结束日期": "期间合计"}
            for c in wide2.columns:
                tot2[c] = round(float(wide2[c].sum()), 2)
            out2 = pd.concat([out2, pd.DataFrame([tot2])], ignore_index=True)

            hdr_row3 = 3
            for c, h in enumerate(list(out2.columns), start=1):
                cell = ws3.cell(row=hdr_row3, column=c, value=str(h))
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal="center")
            for r, (_, row) in enumerate(out2.iterrows(), start=hdr_row3 + 1):
                for c, v in enumerate(row, start=1):
                    ws3.cell(row=r, column=c, value=v)
            for c in range(1, len(out2.columns) + 1):
                ws3.column_dimensions[get_column_letter(c)].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()

    display_name = f"云煤矿业产销量统计_{s0}_{s1}.xlsx"
    ascii_name = f"ymky_viz_{s0}_{s1}.xlsx"
    return content, ascii_name, display_name
