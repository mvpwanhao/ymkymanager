# Copyright (c) 2026-2027 宛皓 (Wan Hao). All rights reserved.
# 本文件为「云煤矿业产销量管理系统」的组成部分。
# 仅授予云南云煤矿业开发有限公司及其关联方在内部业务系统中使用；
# 未经著作权人书面同意，禁止复制、反编译、转售或二次发行。详见根目录 LICENSE。
import os
import shutil
from datetime import date, timedelta

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.config import get_settings
from app.constants import REMOVED_MINE_KEYWORDS
from app.storage import read_records
from app.timeutil import (
    enumerate_weekly_ranges,
    get_26day_month_range,
    get_26day_statistical_month_label,
    get_26day_year_range,
    get_weekly_range,
    today_beijing,
)

EXPORT_RETENTION_DAYS = 7

# sjcl1.xlsx 数据行（A 列）；第 7 行为羊街煤矿，台账已排除，仅占位不填产量
SJCL_V2_DATA_ROWS: tuple[tuple[str, int, str], ...] = (
    ("姚家村", 4, "姚家村煤矿"),
    ("金所", 5, "金所煤矿"),
    ("郭家山", 6, "郭家山煤矿"),
    ("芒东二矿", 8, "芒东二矿"),
    ("胜利", 9, "胜利煤矿"),
    ("竜浪", 10, "竜浪煤矿"),
)
SJCL_V2_YANGJIE_ROW = 7

# weeksheet.xlsx 数据行映射（行号与模板一致）
WEEKSHEET_DATA_ROWS: tuple[tuple[str, int, str], ...] = (
    ("姚家村", 5, "姚家村煤矿"),
    ("金所", 6, "金所煤矿"),
    ("芒东二矿", 7, "芒东二矿"),
    ("郭家山", 8, "郭家山煤矿"),
    ("竜浪", 9, "竜浪煤矿"),
    ("胜利", 10, "胜利煤矿"),
)
WEEKSHEET_TOTAL_ROW = 11


def get_statistical_year_start(target_dt: date) -> date:
    year_start, _ = get_26day_year_range(target_dt)
    return year_start


def set_cell_integer(ws, row, col, value):
    val = float(value) if pd.notna(value) else 0
    cell = ws.cell(row=row, column=col, value=val)
    cell.number_format = "0"


def _cleanup_old_exports(out_dir: str, keep_days: int = EXPORT_RETENTION_DAYS) -> None:
    """删除导出目录中超过保留天数的历史文件（按 mtime）。"""
    try:
        now_ts = today_beijing().timestamp()
    except Exception:
        return
    max_age_seconds = max(int(keep_days), 1) * 24 * 60 * 60
    for name in os.listdir(out_dir):
        p = os.path.join(out_dir, name)
        if not os.path.isfile(p):
            continue
        try:
            st = os.stat(p)
        except OSError:
            continue
        if now_ts - st.st_mtime > max_age_seconds:
            try:
                os.remove(p)
            except OSError:
                # 清理失败不影响主流程
                pass


def _merge_notes(series: pd.Series) -> str:
    if series is None or series.empty:
        return ""
    out: list[str] = []
    seen: set[str] = set()
    for raw in series.tolist():
        s = str(raw or "").strip()
        if not s or s.lower() == "nan":
            continue
        if s in seen:
            continue
        seen.add(s)
        out.append(s)
    return "；".join(out)


def read_sjcl_v2_daily_plans_from_template(template_path: str) -> dict[str, float]:
    """从 sjcl1 模板 B 列读取各矿日计划量（吨）。键为台账煤矿全称，如「姚家村煤矿」。"""
    out: dict[str, float] = {}
    if not template_path or not os.path.isfile(template_path):
        return out
    try:
        wb = load_workbook(template_path, data_only=True)
        ws = wb.active
        for _ledger_prefix, row, canonical in SJCL_V2_DATA_ROWS:
            raw = ws.cell(row=row, column=2).value
            try:
                out[canonical] = float(raw) if raw is not None and str(raw).strip() != "" else 0.0
            except (TypeError, ValueError):
                out[canonical] = 0.0
    except Exception:
        return out
    return out


def generate_sjcl_report(target_date) -> tuple[str | None, str]:
    s = get_settings()
    TEMPLATE_SJCL = s.sjcl_template_v2
    ACTUAL_FILE = s.actual_production_path

    if not os.path.exists(TEMPLATE_SJCL):
        return None, f"未找到实际产量模板：{TEMPLATE_SJCL}"

    try:
        df = read_records(ACTUAL_FILE)
        if df.empty:
            return None, "暂无实际产量数据"
        df["生产日期"] = pd.to_datetime(df["生产日期"]).dt.date
        df["所属煤矿"] = df["所属煤矿"].astype(str)
        df = df[~df["所属煤矿"].str.contains("|".join(REMOVED_MINE_KEYWORDS), na=False)].copy()
        target_dt = pd.to_datetime(target_date).date()

        out_dir = os.path.join(s.data_dir, "exports")
        os.makedirs(out_dir, exist_ok=True)
        _cleanup_old_exports(out_dir)
        today_str = today_beijing().strftime("%m.%d").lstrip("0").replace(".0", ".")
        output_fn = os.path.join(
            out_dir, f"{target_dt.year}年云煤矿业产量进尺统计表（{today_str}）.xlsx"
        )

        shutil.copy(TEMPLATE_SJCL, output_fn)
        wb = load_workbook(output_fn)
        ws = wb.active

        year_start = get_statistical_year_start(target_dt)
        # B 列日计划量以模板为准，不在此重算、不覆盖

        for ledger_prefix, row, canonical in SJCL_V2_DATA_ROWS:
            mine_df = df[df["所属煤矿"].str.startswith(ledger_prefix, na=False)].copy()

            c_val = mine_df[mine_df["生产日期"] == target_dt]["产量(吨)"].sum()
            set_cell_integer(ws, row, 3, c_val)

            mine_day_df = mine_df[mine_df["生产日期"] == target_dt]
            note_text = _merge_notes(mine_day_df["备注"]) if "备注" in mine_day_df.columns else ""
            ws.cell(row=row, column=5, value=note_text or None)

            f_val = mine_df[(mine_df["生产日期"] >= year_start) & (mine_df["生产日期"] <= target_dt)][
                "产量(吨)"
            ].sum()
            set_cell_integer(ws, row, 6, f_val)

        set_cell_integer(ws, SJCL_V2_YANGJIE_ROW, 3, 0)
        ws.cell(row=SJCL_V2_YANGJIE_ROW, column=5, value=None)
        set_cell_integer(ws, SJCL_V2_YANGJIE_ROW, 6, 0)

        for _r in range(4, 11):
            ws.cell(
                row=_r,
                column=4,
                value=f"=IF(B{_r}=0,0,C{_r}/B{_r}*100)",
            )

        r11 = 11
        ws.cell(row=r11, column=2, value="=SUM(B4:B10)").number_format = "0"
        c11 = ws.cell(row=r11, column=3, value="=SUM(C4:C10)")
        c11.number_format = "0"
        ws.cell(row=r11, column=4, value="=IF(B11=0,0,C11/B11*100)")
        ws.cell(row=r11, column=6, value="=SUM(F4:F10)").number_format = "0"

        wb.save(output_fn)
        return output_fn, "实际产量报表生成成功"
    except Exception as e:
        return None, f"出错: {e!s}"


def generate_nybb_report(target_date) -> tuple[str | None, str]:
    s = get_settings()
    TEMPLATE_NYBB = s.nybb_template
    ENERGY_FILE = s.energy_reporting_path

    if not os.path.exists(TEMPLATE_NYBB):
        return None, f"未找到能源局模板：{TEMPLATE_NYBB}"

    try:
        df = read_records(ENERGY_FILE)
        if df.empty:
            return None, "暂无能源局产销量填报数据"
        df["生产日期"] = pd.to_datetime(df["生产日期"]).dt.date
        df["所属煤矿"] = df["所属煤矿"].astype(str)
        df = df[~df["所属煤矿"].str.contains("|".join(REMOVED_MINE_KEYWORDS), na=False)].copy()
        target_dt = pd.to_datetime(target_date).date()

        out_dir = os.path.join(s.data_dir, "exports")
        os.makedirs(out_dir, exist_ok=True)
        _cleanup_old_exports(out_dir)
        today_str = today_beijing().strftime("%m.%d").lstrip("0").replace(".0", ".")
        output_fn = os.path.join(
            out_dir, f"{target_dt.year}年云煤矿业煤炭产量、销量、流向日报表（{today_str}）.xlsx"
        )

        shutil.copy(TEMPLATE_NYBB, output_fn)
        wb = load_workbook(output_fn)
        ws = wb.active

        ws["R2"] = f"日期：{target_dt.strftime('%Y年%m月%d日')}"

        NYBB_MAP = {"郭家山": 5, "姚家村": 6, "金所": 7, "芒东二矿": 8, "胜利": 9, "竜浪": 10}
        current_stat_month = get_26day_statistical_month_label(target_dt)
        year_start = get_statistical_year_start(target_dt)

        for mine, row in NYBB_MAP.items():
            mine_df = df[df["所属煤矿"].str.startswith(mine, na=False)].copy()
            mine_df["统计月"] = mine_df["生产日期"].apply(get_26day_statistical_month_label)

            set_cell_integer(ws, row, 5, mine_df[mine_df["生产日期"] == target_dt]["产量(吨)"].sum())
            set_cell_integer(ws, row, 7, mine_df[mine_df["生产日期"] == target_dt]["销量(吨)"].sum())

            r_val = mine_df[(mine_df["统计月"] == current_stat_month) & (mine_df["生产日期"] <= target_dt)]["产量(吨)"].sum()
            set_cell_integer(ws, row, 18, r_val)

            s_val = mine_df[(mine_df["生产日期"] >= year_start) & (mine_df["生产日期"] <= target_dt)]["产量(吨)"].sum()
            set_cell_integer(ws, row, 19, s_val)

            u_val = mine_df[(mine_df["统计月"] == current_stat_month) & (mine_df["生产日期"] <= target_dt)]["销量(吨)"].sum()
            set_cell_integer(ws, row, 21, u_val)

        sum_cols = [5, 7, 18, 19, 21]
        for c in sum_cols:
            col_let = get_column_letter(c)
            cell = ws.cell(row=11, column=c, value=f"=SUM({col_let}5:{col_let}10)")
            cell.number_format = "0"

        wb.save(output_fn)
        return output_fn, "能源局报表生成成功"
    except Exception as e:
        return None, f"出错: {e!s}"


def _set_cell_num(ws, row, col, value, *, decimals=0):
    """写入数值并设置数字格式。"""
    val = float(value) if pd.notna(value) else 0
    cell = ws.cell(row=row, column=col, value=val)
    cell.number_format = "0" if decimals == 0 else f"0.{'0' * decimals}"


def generate_weekly_report(target_date) -> tuple[str | None, str]:
    """生成周报表（吨表 + 万吨表）。

    数据来源：
    - C/D/E 列（原煤生产量-每周/月累计/年累计）← actual_production 台账按周汇总
    - F/G/H 列（自产煤销售量-每周/月累计/年累计）← actual_sales 台账按周读取
    """
    s = get_settings()
    TEMPLATE = s.weeksheet_template
    ACTUAL_FILE = s.actual_production_path
    SALES_FILE = s.actual_sales_path

    if not os.path.exists(TEMPLATE):
        return None, f"未找到周报表模板：{TEMPLATE}"

    try:
        target_dt = pd.to_datetime(target_date).date()
        week_start, week_end = get_weekly_range(target_dt)
        month_start, month_end = get_26day_month_range(target_dt)
        year_start, _ = get_26day_year_range(target_dt)

        # ── 读取产量台账 ──
        prod_df = read_records(ACTUAL_FILE)
        if not prod_df.empty:
            prod_df["生产日期"] = pd.to_datetime(prod_df["生产日期"]).dt.date
            prod_df["所属煤矿"] = prod_df["所属煤矿"].astype(str)
            prod_df = prod_df[
                ~prod_df["所属煤矿"].str.contains("|".join(REMOVED_MINE_KEYWORDS), na=False)
            ].copy()

        # ── 读取销量台账 ──
        sales_df = read_records(SALES_FILE)
        sales_totals_df = pd.DataFrame()
        if not sales_df.empty:
            sales_df["周起始日期"] = pd.to_datetime(sales_df["周起始日期"], errors="coerce").dt.date
            sales_df["周结束日期"] = pd.to_datetime(sales_df["周结束日期"], errors="coerce").dt.date
            sales_df["所属煤矿"] = sales_df["所属煤矿"].astype(str)
            sales_df["销量(吨)"] = pd.to_numeric(sales_df["销量(吨)"], errors="coerce").fillna(0)
            for _col in [
                "月累计自产煤销量(吨)", "年累计自产煤销量(吨)",
                "年累计掺配煤销量(吨)", "年累计外购煤量(吨)",
            ]:
                if _col not in sales_df.columns:
                    sales_df[_col] = 0.0
                sales_df[_col] = pd.to_numeric(sales_df[_col], errors="coerce").fillna(0)
            # 分离"合计"记录（公司级 I/J 数据）与矿记录
            total_mask = sales_df["所属煤矿"] == "合计"
            sales_totals_df = sales_df[total_mask].copy()
            sales_df = sales_df[~total_mask].copy()
            sales_df = sales_df[
                ~sales_df["所属煤矿"].str.contains("|".join(REMOVED_MINE_KEYWORDS), na=False)
            ].copy()

        out_dir = os.path.join(s.data_dir, "exports")
        os.makedirs(out_dir, exist_ok=True)
        _cleanup_old_exports(out_dir)
        _date_cn = lambda d: f"{d.month}月{d.day}日"
        output_fn = os.path.join(
            out_dir,
            f"每周生产、销售煤量统计表（{_date_cn(week_start)}-{_date_cn(week_end)}）.xlsx",
        )

        shutil.copy(TEMPLATE, output_fn)
        wb = load_workbook(output_fn)

        # ── 计算各矿数据 ──
        mine_data: dict[str, dict[str, float]] = {}
        for ledger_prefix, row, canonical in WEEKSHEET_DATA_ROWS:
            md: dict[str, float] = {}

            # 产量：C(每周), D(月累计), E(年累计)
            if prod_df.empty:
                md["c"] = md["d"] = md["e"] = 0.0
            else:
                mdf = prod_df[prod_df["所属煤矿"].str.startswith(ledger_prefix, na=False)]
                md["c"] = mdf[(mdf["生产日期"] >= week_start) & (mdf["生产日期"] <= week_end)]["产量(吨)"].sum()
                md["d"] = mdf[(mdf["生产日期"] >= month_start) & (mdf["生产日期"] <= week_end)]["产量(吨)"].sum()
                md["e"] = mdf[(mdf["生产日期"] >= year_start) & (mdf["生产日期"] <= week_end)]["产量(吨)"].sum()

            # 销量：F(每周), G(月累计), H(年累计)
            # I/J 为公司级数据，不写入各矿行——仅从"合计"记录取 I11/J11
            if sales_df.empty:
                md["f"] = md["g"] = md["h"] = md["i"] = md["j"] = 0.0
            else:
                sdf = sales_df[sales_df["所属煤矿"].str.startswith(ledger_prefix, na=False)]
                md["f"] = sdf[(sdf["周起始日期"] == week_start) & (sdf["周结束日期"] == week_end)]["销量(吨)"].sum()
                # G/H: 以补录存储值为基数 + 后续新增F值；无存储值时按时间累计
                def _calc_cumul(_sdf, _period_start, _cumul_col):
                    _recs = _sdf[(_sdf["周结束日期"] >= _period_start) & (_sdf["周结束日期"] <= week_end)]
                    if _recs.empty:
                        return 0.0
                    _stored = _recs[_recs[_cumul_col] > 0]
                    if not _stored.empty:
                        _latest = _stored.sort_values("周结束日期").iloc[-1]
                        _base = float(_latest[_cumul_col])
                        _base_we = _latest["周结束日期"]
                        _add = _recs[_recs["周结束日期"] > _base_we]["销量(吨)"].sum()
                        return _base + float(_add)
                    return float(_recs["销量(吨)"].sum())

                md["g"] = _calc_cumul(sdf, month_start, "月累计自产煤销量(吨)")
                md["h"] = _calc_cumul(sdf, year_start, "年累计自产煤销量(吨)")
                # I/J 各矿行不写入
                md["i"] = 0.0
                md["j"] = 0.0

            mine_data[canonical] = md

        # ── 日期范围字符串 ──
        def _cn_date(d: date) -> str:
            return f"{d.year}年{d.month}月{d.day}日"

        date_range_str = f"{_cn_date(week_start)}-{_cn_date(week_end)}"

        # ── 填充吨表 ──
        ws_ton = wb["吨表"]
        ws_ton["A2"] = f"{date_range_str}                                                                                             单位：吨"
        for ledger_prefix, row, canonical in WEEKSHEET_DATA_ROWS:
            md = mine_data[canonical]
            _set_cell_num(ws_ton, row, 3, md["c"], decimals=2)   # C: 每周产量
            _set_cell_num(ws_ton, row, 4, md["d"], decimals=2)   # D: 月累计产量
            _set_cell_num(ws_ton, row, 5, md["e"], decimals=2)   # E: 年累计产量
            _set_cell_num(ws_ton, row, 6, md["f"], decimals=2)   # F: 每周销量
            _set_cell_num(ws_ton, row, 7, md["g"], decimals=2)   # G: 月累计销量
            _set_cell_num(ws_ton, row, 8, md["h"], decimals=2)   # H: 年累计销量
            _set_cell_num(ws_ton, row, 9, md["i"], decimals=2)   # I: 年累计掺配煤
            _set_cell_num(ws_ton, row, 10, md["j"], decimals=2)  # J: 年累计外购煤

        # 吨表合计行：C-H 列直接计算写入；I11/J11 从"合计"记录取（无则求和各矿）
        for col in range(3, 9):
            total = sum(
                (ws_ton.cell(row=r, column=col).value or 0) for r in range(5, 11)
            )
            _set_cell_num(ws_ton, WEEKSHEET_TOTAL_ROW, col, total, decimals=2)
        # I/J 各矿行为 0，合计行从"合计"记录取（无则回退到最近一期合计记录）
        _i11 = 0.0
        _j11 = 0.0
        if not sales_totals_df.empty:
            tot_week = sales_totals_df[sales_totals_df["周结束日期"] == week_end]
            if tot_week.empty:
                _recent_tot = sales_totals_df[sales_totals_df["周结束日期"] <= week_end]
                if not _recent_tot.empty:
                    tot_week = _recent_tot.sort_values("周结束日期").iloc[[-1]]
            if not tot_week.empty:
                _i11 = float(tot_week["年累计掺配煤销量(吨)"].iloc[0])
                _j11 = float(tot_week["年累计外购煤量(吨)"].iloc[0])
        if _i11 == 0.0:
            _i11 = sum((ws_ton.cell(row=r, column=9).value or 0) for r in range(5, 11))
        if _j11 == 0.0:
            _j11 = sum((ws_ton.cell(row=r, column=10).value or 0) for r in range(5, 11))
        _set_cell_num(ws_ton, WEEKSHEET_TOTAL_ROW, 9, _i11, decimals=2)
        _set_cell_num(ws_ton, WEEKSHEET_TOTAL_ROW, 10, _j11, decimals=2)
        h_total = ws_ton.cell(row=WEEKSHEET_TOTAL_ROW, column=8).value or 0
        _set_cell_num(ws_ton, WEEKSHEET_TOTAL_ROW, 11, h_total + _i11, decimals=2)  # K = H + I

        # ── 填充万吨表 ──
        ws_wan = wb["万吨表"]
        ws_wan["A2"] = f"{date_range_str}                                                                                             单位：万吨"
        for ledger_prefix, row, canonical in WEEKSHEET_DATA_ROWS:
            md = mine_data[canonical]
            _set_cell_num(ws_wan, row, 3, md["c"] / 10000, decimals=1)
            _set_cell_num(ws_wan, row, 4, md["d"] / 10000, decimals=1)
            _set_cell_num(ws_wan, row, 5, md["e"] / 10000, decimals=1)
            _set_cell_num(ws_wan, row, 6, md["f"] / 10000, decimals=1)
            _set_cell_num(ws_wan, row, 7, md["g"] / 10000, decimals=1)
            _set_cell_num(ws_wan, row, 8, md["h"] / 10000, decimals=1)
            _set_cell_num(ws_wan, row, 9, md["i"] / 10000, decimals=1)
            _set_cell_num(ws_wan, row, 10, md["j"] / 10000, decimals=1)

        # 万吨表合计行：C-H 列直接计算写入；I11/J11 从"合计"记录取
        for col in range(3, 9):
            total = sum(
                (ws_wan.cell(row=r, column=col).value or 0) for r in range(5, 11)
            )
            _set_cell_num(ws_wan, WEEKSHEET_TOTAL_ROW, col, total, decimals=1)
        _set_cell_num(ws_wan, WEEKSHEET_TOTAL_ROW, 9, _i11 / 10000, decimals=1)
        _set_cell_num(ws_wan, WEEKSHEET_TOTAL_ROW, 10, _j11 / 10000, decimals=1)
        h_total_w = ws_wan.cell(row=WEEKSHEET_TOTAL_ROW, column=8).value or 0
        _set_cell_num(ws_wan, WEEKSHEET_TOTAL_ROW, 11, h_total_w + _i11 / 10000, decimals=1)  # K = H + I

        wb.save(output_fn)
        return output_fn, f"周报表生成成功（{date_range_str}）"
    except Exception as e:
        return None, f"出错: {e!s}"


# 产销量简报中各矿排列顺序
BRIEF_MINE_ORDER: tuple[str, ...] = (
    "姚家村",
    "金所",
    "芒东二矿",
    "郭家山",
    "竜浪",
    "胜利",
)


def generate_brief_report(target_date) -> tuple[str | None, str]:
    """生成产销量简报文本（纯文本，可直接复制粘贴到微信群等）。

    返回 (brief_text, message)；brief_text 为 None 表示出错。
    """
    s = get_settings()
    ACTUAL_FILE = s.actual_production_path
    SALES_FILE = s.actual_sales_path

    try:
        target_dt = pd.to_datetime(target_date).date()
        week_start, week_end = get_weekly_range(target_dt)
        month_start, month_end = get_26day_month_range(target_dt)
        year_start, _ = get_26day_year_range(target_dt)

        # ── 确定周序号 ──
        all_weeks = enumerate_weekly_ranges(month_start, month_end)
        week_num = 1
        for i, (ws, we) in enumerate(all_weeks, 1):
            if ws == week_start and we == week_end:
                week_num = i
                break

        # ── 上周区间 ──
        prev_week_start = week_start - timedelta(days=7)
        prev_week_end = week_end - timedelta(days=7)

        # ── 读取产量台账 ──
        prod_df = read_records(ACTUAL_FILE)
        if not prod_df.empty:
            prod_df["生产日期"] = pd.to_datetime(prod_df["生产日期"]).dt.date
            prod_df["所属煤矿"] = prod_df["所属煤矿"].astype(str)
            prod_df = prod_df[
                ~prod_df["所属煤矿"].str.contains("|".join(REMOVED_MINE_KEYWORDS), na=False)
            ].copy()
            prod_df["产量(吨)"] = pd.to_numeric(prod_df["产量(吨)"], errors="coerce").fillna(0)

        # ── 读取销量台账 ──
        sales_df = read_records(SALES_FILE)
        sales_totals_df = pd.DataFrame()
        if not sales_df.empty:
            sales_df["周起始日期"] = pd.to_datetime(sales_df["周起始日期"], errors="coerce").dt.date
            sales_df["周结束日期"] = pd.to_datetime(sales_df["周结束日期"], errors="coerce").dt.date
            sales_df["所属煤矿"] = sales_df["所属煤矿"].astype(str)
            sales_df["销量(吨)"] = pd.to_numeric(sales_df["销量(吨)"], errors="coerce").fillna(0)
            # 新增累计列
            for _col in [
                "月累计自产煤销量(吨)", "年累计自产煤销量(吨)",
                "年累计掺配煤销量(吨)", "年累计外购煤量(吨)",
            ]:
                if _col not in sales_df.columns:
                    sales_df[_col] = 0.0
                sales_df[_col] = pd.to_numeric(sales_df[_col], errors="coerce").fillna(0)
            # 分离"合计"记录（公司级 I/J 数据）与矿记录
            total_mask = sales_df["所属煤矿"] == "合计"
            sales_totals_df = sales_df[total_mask].copy()
            sales_df = sales_df[~total_mask].copy()
            sales_df = sales_df[
                ~sales_df["所属煤矿"].str.contains("|".join(REMOVED_MINE_KEYWORDS), na=False)
            ].copy()

        def _sum_prod(d_start: date, d_end: date) -> float:
            if prod_df.empty:
                return 0.0
            mask = (prod_df["生产日期"] >= d_start) & (prod_df["生产日期"] <= d_end)
            return float(prod_df.loc[mask, "产量(吨)"].sum())

        def _sum_sales(d_start: date, d_end: date) -> float:
            """按周结束日期筛选销量（销量台账为周频）。"""
            if sales_df.empty:
                return 0.0
            mask = (sales_df["周结束日期"] >= d_start) & (sales_df["周结束日期"] <= d_end)
            return float(sales_df.loc[mask, "销量(吨)"].sum())

        # ── 计算各项数据 ──
        week_prod = _sum_prod(week_start, week_end)
        prev_week_prod = _sum_prod(prev_week_start, prev_week_end)
        month_prod = _sum_prod(month_start, week_end)
        year_prod = _sum_prod(year_start, week_end)
        wow_delta = week_prod - prev_week_prod

        week_sales = _sum_sales(week_start, week_end)

        # 确定有效数据周：优先用目标周，无数据时回退到最近一期（<= week_end）
        effective_week_end = week_end
        if not sales_df.empty:
            exact_week = sales_df[sales_df["周结束日期"] == week_end]
            if exact_week.empty:
                recent = sales_df[sales_df["周结束日期"] <= week_end]
                if not recent.empty:
                    effective_week_end = recent["周结束日期"].max()

        # 月累计/年累计自产煤销量：以"合计"记录存储G/H为基数 + 后续新增F值；无存储值时按时间累计
        def _calc_cumul_brief(_period_start, _cumul_col):
            if not sales_totals_df.empty:
                _period_tot = sales_totals_df[
                    (sales_totals_df["周结束日期"] >= _period_start) &
                    (sales_totals_df["周结束日期"] <= week_end)
                ]
                if not _period_tot.empty:
                    _stored = _period_tot[_period_tot[_cumul_col] > 0]
                    if not _stored.empty:
                        _latest = _stored.sort_values("周结束日期").iloc[-1]
                        _base = float(_latest[_cumul_col])
                        _base_we = _latest["周结束日期"]
                        _add = sales_df[
                            (sales_df["周结束日期"] > _base_we) &
                            (sales_df["周结束日期"] <= week_end)
                        ]["销量(吨)"].sum()
                        return _base + float(_add)
            return _sum_sales(_period_start, week_end)

        month_sales = _calc_cumul_brief(month_start, "月累计自产煤销量(吨)")
        year_sales = _calc_cumul_brief(year_start, "年累计自产煤销量(吨)")

        # 掺配煤、外购煤：优先从"合计"记录取，无则从各矿记录汇总
        # "合计"记录使用独立的有效周回退（与矿记录分开）
        effective_week_end_totals = week_end
        if not sales_totals_df.empty:
            _exact_tot = sales_totals_df[sales_totals_df["周结束日期"] == week_end]
            if _exact_tot.empty:
                _recent_tot = sales_totals_df[sales_totals_df["周结束日期"] <= week_end]
                if not _recent_tot.empty:
                    effective_week_end_totals = _recent_tot["周结束日期"].max()

        blended_sales = 0.0
        purchased_coal = 0.0
        if not sales_totals_df.empty:
            tot_week = sales_totals_df[sales_totals_df["周结束日期"] == effective_week_end_totals]
            if not tot_week.empty:
                blended_sales = float(tot_week["年累计掺配煤销量(吨)"].iloc[0])
                purchased_coal = float(tot_week["年累计外购煤量(吨)"].iloc[0])
        if blended_sales == 0.0 and not sales_df.empty:
            week_mines_ij = sales_df[sales_df["周结束日期"] == effective_week_end]
            if not week_mines_ij.empty:
                blended_sales = float(week_mines_ij["年累计掺配煤销量(吨)"].sum())
        if purchased_coal == 0.0 and not sales_df.empty:
            week_mines_ij = sales_df[sales_df["周结束日期"] == effective_week_end]
            if not week_mines_ij.empty:
                purchased_coal = float(week_mines_ij["年累计外购煤量(吨)"].sum())
        # K = H + I（合计销售煤量 = 自产煤年累计 + 掺配煤年累计）
        total_sales = year_sales + blended_sales

        # ── 各矿年累计产量（万吨）──
        mine_yearly: list[tuple[str, float]] = []
        for prefix in BRIEF_MINE_ORDER:
            if prod_df.empty:
                mine_yearly.append((prefix, 0.0))
                continue
            mdf = prod_df[prod_df["所属煤矿"].str.startswith(prefix, na=False)]
            val = float(mdf.loc[
                (mdf["生产日期"] >= year_start) & (mdf["生产日期"] <= week_end),
                "产量(吨)",
            ].sum())
            mine_yearly.append((prefix, val))

        # ── 日期格式化辅助 ──
        stat_month = month_end.month  # 统计月 = 25日所在月
        year_val = target_dt.year

        def _md(d: date) -> str:
            return f"{d.month}月{d.day}日"

        # ── 组装简报文本 ──
        lines: list[str] = []
        lines.append(f"云煤矿业公司{year_val}年生产销售情况汇报：")
        lines.append(
            f"{stat_month}月第{week_num}周（{_md(week_start)}至{_md(week_end)}）生产量：{week_prod:,.0f}吨"
        )
        lines.append(
            f"{stat_month}月（{_md(month_start)}至{_md(week_end)}）累计生产量：{month_prod:,.0f}吨"
        )
        lines.append(f"年累计总生产量：{year_prod:,.0f}吨")
        wow_str = f"{wow_delta:+,.0f}" if wow_delta != 0 else "0"
        lines.append(f"环比上周增量：{wow_str}吨")
        lines.append("－－－－－－－－－－－－－")
        lines.append(
            f"{stat_month}月第{week_num}周（{_md(week_start)}至{_md(week_end)}）自产煤销售量：{week_sales:,.0f}吨"
        )
        lines.append(f"{stat_month}月累计自产煤销售量：{month_sales:,.0f}吨")
        lines.append(f"年累计自产煤销售量：{year_sales:,.0f}吨")
        lines.append(f"年累计掺配煤销售量：{blended_sales:,.0f}吨")
        lines.append(f"年累计外购煤量：{purchased_coal:,.0f}吨")
        lines.append(f"年合计销售煤量：{total_sales:,.0f}吨")
        lines.append("－－－－－－－－－－－－－－")
        lines.append(f"{year_val}年累计生产量：")
        mine_labels = {
            "姚家村": "姚家村煤矿",
            "金所": "金所煤矿",
            "芒东二矿": "芒东二矿",
            "郭家山": "郭家山煤矿",
            "竜浪": "竜浪煤矿",
            "胜利": "胜利煤矿",
        }
        for i, (prefix, val) in enumerate(mine_yearly, 1):
            label = mine_labels.get(prefix, prefix)
            lines.append(f"   {i}、{label} {val / 10000:,.1f}万吨；")

        brief_text = "\n".join(lines)
        return brief_text, "产销量简报生成成功"
    except Exception as e:
        return None, f"出错: {e!s}"
