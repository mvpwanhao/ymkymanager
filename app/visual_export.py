# Copyright (c) 2026-2027 宛皓 (Wan Hao). All rights reserved.
# 本文件为「云煤矿业产销量管理系统」的组成部分。
# 仅授予云南云煤矿业开发有限公司及其关联方在内部业务系统中使用；
# 未经著作权人书面同意，禁止复制、反编译、转售或二次发行。详见根目录 LICENSE。
"""数据可视化导出：日均产量明细表 + Plotly 三图 PNG 写入 xlsx。"""

from __future__ import annotations

import io
import os
import re
from datetime import date
from typing import Any
from urllib.parse import quote

import plotly.graph_objects as go
import plotly.io as pio
from fastapi import Request
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

from app.dashboard_data import build_daily_mines_export_table, build_summary_and_charts
from app.timeutil import today_beijing


def visual_build_kwargs_from_request(request: Request) -> dict[str, Any]:
    """与 GET / 可视化页一致的 query 解析，传给 build_summary_and_charts。"""
    period = request.query_params.get("period", "year")
    ds = request.query_params.get("start")
    de = request.query_params.get("end")
    today_cap = today_beijing()
    c_start: date | None = None
    c_end: date | None = None
    tpl_start = (ds or "").strip()
    tpl_end = (de or "").strip()
    if period == "custom" and tpl_start and tpl_end:
        try:
            c_start = date.fromisoformat(str(tpl_start))
            c_end = date.fromisoformat(str(tpl_end))
            if c_start > today_cap:
                c_start = today_cap
            if c_end > today_cap:
                c_end = today_cap
            if c_start > c_end:
                c_start, c_end = c_end, c_start
        except ValueError:
            period = "year"
    sy_raw = (request.query_params.get("stat_year") or "").strip()
    sm_raw = (request.query_params.get("stat_month") or "").strip()
    stat_year_int: int | None = int(sy_raw) if sy_raw.isdigit() else None
    stat_month_str: str | None = sm_raw if re.fullmatch(r"\d{4}-\d{2}", sm_raw) else None
    return {
        "period": period,
        "custom_start": c_start,
        "custom_end": c_end,
        "stat_year": stat_year_int,
        "stat_month": stat_month_str,
    }


def _prepare_kaleido_export_env() -> None:
    """Kaleido 子进程写 fontconfig 等缓存；确保可写目录存在，避免选字回退异常。"""
    base = "/tmp/.ymky-kaleido-cache"
    os.makedirs(base, mode=0o700, exist_ok=True)
    os.environ.setdefault("XDG_CACHE_HOME", base)


def _fig_to_png_bytes(fig: go.Figure, width: int = 980, height: int = 520) -> bytes:
    try:
        _prepare_kaleido_export_env()
        return pio.to_image(fig, format="png", width=width, height=height, scale=2, engine="kaleido")
    except Exception as exc:
        raise RuntimeError(
            "导出图表快照失败（Plotly + Kaleido）。请确认 requirements 中 kaleido<1"
            "（1.x 依赖系统 Chrome/choreographer）；重建镜像后 docker compose up -d --force-recreate ymky。"
            "详见容器日志中本次异常链。"
        ) from exc


def _scale_xl_image(img: XLImage, max_w_px: int = 920) -> None:
    w = getattr(img, "width", None)
    if not w or w <= max_w_px:
        return
    ratio = max_w_px / float(w)
    img.width = int(max_w_px)
    h = getattr(img, "height", None)
    if h:
        img.height = int(h * ratio)


def try_visual_export_bytes(
    request: Request,
) -> tuple[bytes | None, str | None, str | None, str | None]:
    """
    返回 (xlsx_bytes, ascii_filename, utf8_filename, error_message)。
    error 非空表示失败。
    """
    kw = visual_build_kwargs_from_request(request)
    d = build_summary_and_charts(include_export_figures=True, **kw)
    if d.get("empty"):
        return None, None, None, str(d.get("message") or "暂无可导出数据")
    bag = d.get("_export")
    if not isinstance(bag, dict):
        return None, None, None, "当前区间暂无产量数据，无法导出。"
    pstart = bag.get("period_start")
    pend = bag.get("period_end")
    pdf = bag["period_df"]
    pie_f: go.Figure = bag["pie_figure"]
    bar_f: go.Figure = bag["bar_figure"]
    tr_f: go.Figure = bag["trend_figure"]
    prange = str(d.get("period_caption") or "")

    table = build_daily_mines_export_table(pdf)

    try:
        png_pie = _fig_to_png_bytes(pie_f, 980, 520)
        png_bar = _fig_to_png_bytes(bar_f, 980, 480)
        png_tr = _fig_to_png_bytes(tr_f, 1000, 540)

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "日均产量"
        ws1["A1"] = "产量统计导出"
        ws1["A1"].font = Font(bold=True, size=14)
        ws1["A2"] = f"统计区间：{prange}"
        ws1["A2"].font = Font(bold=False, size=11)
        header_row = 4
        for r_i, row in enumerate(dataframe_to_rows(table, index=False, header=True), start=header_row):
            for c_i, cell in enumerate(row, start=1):
                ws1.cell(row=r_i, column=c_i, value=cell)

        ws2 = wb.create_sheet("图表", 1)
        row_cursor = 1
        bundles = [
            ("各矿产量占比（饼状图）", png_pie),
            ("各矿产量柱状图", png_bar),
            ("日产量趋势（折线图）", png_tr),
        ]
        for title, pngb in bundles:
            cell = ws2.cell(row=row_cursor, column=1, value=title)
            cell.font = Font(bold=True, size=12)
            img = XLImage(io.BytesIO(pngb))
            _scale_xl_image(img)
            ws2.add_image(img, f"A{row_cursor + 1}")
            row_cursor += max(35, int((getattr(img, "height", 420) or 420) / 18) + 5)

        buf = io.BytesIO()
        wb.save(buf)
        content = buf.getvalue()
    except RuntimeError as e:
        return None, None, None, str(e)

    s0 = pstart.isoformat() if hasattr(pstart, "isoformat") else str(pstart)
    s1 = pend.isoformat() if hasattr(pend, "isoformat") else str(pend)
    display_name = f"云煤矿业产量可视化_{s0}_{s1}.xlsx"
    ascii_name = f"ymky_visual_{s0}_{s1}.xlsx"
    return content, ascii_name, display_name, None


def content_disposition_attachment(ascii_name: str, utf8_name: str) -> str:
    return f'attachment; filename="{ascii_name}"; filename*=UTF-8\'\'{quote(utf8_name)}'
